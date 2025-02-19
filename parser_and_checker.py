import os
import xlsxwriter
from openpyxl import load_workbook
import subprocess
import re

def split_cell(cell):
    match = re.match(r"([A-Z]+)(\d+)", cell)
    if match:
        return match.groups()
    else:
        raise ValueError(f"Invalid cell format: {cell}")

MAX_ROWS = 1000
MAX_COLS = 1000

def format_range(range_str):
    """Format range for Excel formulas."""
    start, end = range_str.split(':')
    return f"{start}:{end}"

def process_testcase(workbook, test_number, operations):
    """Initialize worksheet with 0 and write operations."""
    worksheet = workbook.add_worksheet(f'Test{test_number}')

    # Step 1: Initialize entire worksheet with 0
    for row in range(MAX_ROWS):
        for col in range(MAX_COLS):
            worksheet.write_number(row, col, 0)

    # Step 2: Process operations
    for operation in operations:
        if '=' not in operation:
            continue

        cell, formula = operation.split('=')
        col_str, row_str = split_cell(cell)
        col = sum((ord(char) - ord('A') + 1) * (26 ** i) for i, char in enumerate(reversed(col_str))) - 1
        row = int(row_str) - 1

        if 'SLEEP' in formula:
            sleep_value = int(float(formula[formula.find('(') + 1:formula.find(')')]))
            worksheet.write_number(row, col, sleep_value)
        elif 'SUM' in formula:
            excel_range = format_range(formula[formula.find('(') + 1:formula.find(')')])
            worksheet.write_formula(row, col, f'=TRUNC(SUM({excel_range}))')
        elif 'AVG' in formula:
            excel_range = format_range(formula[formula.find('(') + 1:formula.find(')')])
            worksheet.write_formula(row, col, f'=TRUNC(AVERAGE({excel_range}))')
        elif 'STDEV' in formula:
            excel_range = format_range(formula[formula.find('(') + 1:formula.find(')')])
            worksheet.write_formula(row, col, f'=TRUNC(SQRT((SUMPRODUCT({excel_range}, {excel_range})/COUNT({excel_range})) - (AVERAGE({excel_range})^2)))')
        elif 'MIN' in formula:
            excel_range = format_range(formula[formula.find('(') + 1:formula.find(')')])
            worksheet.write_formula(row, col, f'=MIN({excel_range})')
        elif 'MAX' in formula:
            excel_range = format_range(formula[formula.find('(') + 1:formula.find(')')])
            worksheet.write_formula(row, col, f'=MAX({excel_range})')
        else:
            try:
                value = float(formula)
                worksheet.write_number(row, col, value)
            except ValueError:
                if formula[0] != '=':
                    formula = f'=TRUNC({formula})'
                worksheet.write_formula(row, col, formula)

def force_excel_recalculation(file_path):
    """Use LibreOffice to recalculate formulas on Linux."""
    if not os.path.exists(file_path):
        print(f"Error: {file_path} not found.")
        return

    try:
        # Convert to ods and back to xlsx to force recalculation
        subprocess.run(['soffice', '--headless', '--convert-to', 'ods', file_path],
                       check=True, capture_output=True)
        ods_file = file_path.replace('.xlsx', '.ods')
        subprocess.run(['soffice', '--headless', '--convert-to', 'xlsx', ods_file],
                       check=True, capture_output=True)

        # Clean up temporary ods file
        if os.path.exists(ods_file):
            os.remove(ods_file)
    except subprocess.CalledProcessError as e:
        print(f"Error during LibreOffice conversion: {e}")
    except Exception as e:
        print(f"Unexpected error during recalculation: {e}")

def main():
    workbook_filename = 'temp.xlsx'

    # Create workbook
    workbook = xlsxwriter.Workbook(workbook_filename)

    # Read input
    with open('tc1.txt', 'r') as f:
        lines = f.readlines()

    test_cases = int(lines[0])
    current_line = 1
    query_info = []

    for test_num in range(test_cases):
        n, m = map(int, lines[current_line].split())
        operations = [lines[current_line + i].strip() for i in range(1, n + 1)]
        queries = [lines[current_line + n + i].strip() for i in range(1, m + 1)]

        process_testcase(workbook, test_num + 1, operations)

        for query in queries:
            query_info.append((test_num + 1, query))

        current_line += n + m + 1

    workbook.close()

    # Force recalculation using LibreOffice
    force_excel_recalculation(workbook_filename)

    # Read computed values
    wb = load_workbook(workbook_filename, data_only=True)
    output_results = []

    for test_num, query in query_info:
        worksheet = wb[f'Test{test_num}']
        try:
            cell = query.split('=')[0].strip()
            col = ord(cell[0]) - ord('A')
            row = int(cell[1:]) - 1

            value = worksheet.cell(row=row + 1, column=col + 1).value

            if value is None:
                formatted_value = '0'
            elif isinstance(value, (int, float)):
                formatted_value = '{:.6f}'.format(float(value)).rstrip('0').rstrip('.')
            else:
                formatted_value = str(value)

            output_results.append(f"{cell}: {formatted_value}")

        except Exception:
            output_results.append(f"{cell}: 0")

    with open('model_output.txt', 'w') as f:
        f.write('\n'.join(output_results))

if __name__ == "__main__":
    main()