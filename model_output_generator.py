import os
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from pycel import ExcelCompiler
from pycel.excelutil import DIV0

def excel_col_from_index(index):
    if index < 0:
        raise ValueError("Index must be positive integer")
    result = ""
    while index >= 0:
        index, rem = divmod(index, 26)
        result = chr(65 + rem) + result
        index -= 1
    return result

def create_excel_file(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(filename)
    return filename

def cell_to_indices(cell_ref):
    col_str = ''.join(c for c in cell_ref if c.isalpha())
    row_str = ''.join(c for c in cell_ref if c.isdigit())
    col_index = 0
    for char in col_str:
        col_index = col_index * 26 + (ord(char) - ord('A') + 1)
    row_index = int(row_str)
    return row_index, col_index

def load_and_evaluate_testcase(filename, fn):
    orig = filename
    filename = f"testcases/{filename}{fn:02d}.txt"
    xlsx_file = create_excel_file(f"expected_outputs/sheets/{orig}{fn:02d}.xlsx")
    print("Excel File created")
    expected_output = ""
    with open(filename, 'r') as f:
        t = int(f.readline())
        n, m = map(int, f.readline().split())
        formulas = [f.readline().strip() for _ in range(n)]
        queries = [f.readline().strip() for _ in range(m)]

    formula_history = {}
    print("Loading formulas and queries...")
    for _ in range(t):
        for formula in formulas:
            parts = formula.split('=', 1)
            cell = parts[0].strip()
            value = parts[1].strip()
            if 'AVG' in value:
                range_str = value[value.find('(')+1:value.find(')')]
                start, end = range_str.split(':')
                start_row, start_col = cell_to_indices(start)
                end_row, end_col = cell_to_indices(end)
                num_rows = end_row - start_row + 1
                num_cols = end_col - start_col + 1
                num_cells = num_rows * num_cols
                value = f"SUM({start}:{end})/{num_cells}"
            if 'STDEV' in value:
                range_str = value[value.find('(')+1:value.find(')')]
                start, end = range_str.split(':')
                start_row, start_col = cell_to_indices(start)
                end_row, end_col = cell_to_indices(end)
                num_rows = end_row - start_row + 1
                num_cols = end_col - start_col + 1
                num_cells = num_rows * num_cols
                value = f'ROUND(SQRT(SUMPRODUCT((IF(ISBLANK({range_str}),0,{range_str})-AVERAGE(IF(ISBLANK({range_str}),0,{range_str})))^2)/{num_cells}),0)'
            if 'MIN' in value:
                range_str = value[value.find('(')+1:value.find(')')]
                value = f'MIN(IF(ISBLANK({range_str}),0,{range_str}))'
            if 'SLEEP' in value:
                value = value.split('(')[1].split(')')[0]
            if value.isdigit():
                value = int(value)
            else:
                value = f"=TRUNC({value})"
            wb = load_workbook(xlsx_file)
            ws = wb['Sheet1']
            prev_formula = None
            if cell in formula_history:
                prev_formula = formula_history[cell]
            ws[cell] = value
            wb.save(xlsx_file)
            formula_history[cell] = ws[cell].value
            excel = ExcelCompiler(xlsx_file)
            try:
                excel.evaluate(f'Sheet1!{cell}')
                expected_output += f'ok\n'
            except RecursionError:
                expected_output += f'circ\n'
                wb = load_workbook(xlsx_file)
                ws = wb['Sheet1']
                ws[cell] = prev_formula
                formula_history[cell] = prev_formula
                wb.save(xlsx_file)
            print(f"Ran formula: {cell} = {value}")
        excel = ExcelCompiler(xlsx_file)
        for query in queries:
            cell = query.strip()
            try:
                result = excel.evaluate(f'Sheet1!{cell}')
                if result == DIV0:
                    expected_output += f'{cell}: #DIV/0!\n'
                else:
                    expected_output += f'{cell}: {0 if result is None else int(result)}\n'
            except Exception as e:
                expected_output += f'{cell}: {str(e)}\n'

    with open(f'expected_outputs/text_files/{orig}{fn:02d}.txt', 'w') as f:
        f.write(expected_output)

if __name__ == "__main__":
    # iterate through files in testcases directory
    for filename in sorted(os.listdir('testcases')):
        if filename.endswith('_12.txt'):
            load_and_evaluate_testcase(filename.split('_')[0] + '_', int(filename.split('_')[-1].split('.')[0]))